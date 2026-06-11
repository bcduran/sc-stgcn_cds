"""
src/pipeline/04_build_sc_relations.py
======================================
Amac:
  Top-50 sirketi supply chain grafinda ara.
  Graf sirket adlariyla dolu (ticker degil).
  CDS dosyasindan ticker->sirket adi alinir,
  fuzzy matching ile graf dugumlerine eslenir.

Girdi:
  data/processed/raw_edges.csv
  data/processed/e_top50_companies.xlsx
  Data/raw/cds.csv

Cikti:
  data/processed/e_top_50_supplier_info.xlsx
  data/processed/e_top_50_customer_info.xlsx
"""

from __future__ import annotations
from pathlib import Path
import sys, re
from difflib import SequenceMatcher
import pandas as pd
import networkx as nx
from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
from configs.config import (
    CDS_RAW_CSV, TOP50_XLSX, RAW_EDGES_CSV,
    TOP50_SUPPLIER_XLSX, TOP50_CUSTOMER_XLSX, make_dirs,
)

FUZZY_THRESHOLD = 0.72


def _normalize(name: str) -> str:
    s = name.lower().strip()
    s = re.sub(r"[&.,'\-\(\)]", " ", s)
    stopwords = {"inc","corp","co","ltd","llc","plc","the","group",
                 "company","companies","holdings","international",
                 "corporation","incorporated","limited"}
    tokens = [t for t in s.split() if t and t not in stopwords]
    return " ".join(tokens)


def _sim(a: str, b: str) -> float:
    return SequenceMatcher(None, _normalize(a), _normalize(b)).ratio()


def load_graph() -> nx.DiGraph:
    df = pd.read_csv(RAW_EDGES_CSV)
    df.columns = [c.upper() for c in df.columns]
    G = nx.DiGraph()
    G.add_edges_from(
        zip(df["SUPPLIER"].astype(str).str.strip(),
            df["CUSTOMER"].astype(str).str.strip())
    )
    print(f"  Graf: {G.number_of_nodes():,} dugum, {G.number_of_edges():,} kenar")
    return G


def load_top50() -> pd.DataFrame:
    top50 = pd.read_excel(TOP50_XLSX)
    top50["Ticker"] = top50["Ticker"].astype(str).str.strip().str.upper()
    cds = pd.read_csv(CDS_RAW_CSV, usecols=["Ticker","Company"]).drop_duplicates()
    cds["Ticker"] = cds["Ticker"].astype(str).str.strip().str.upper()
    cds_dict = dict(zip(cds["Ticker"], cds["Company"]))
    top50["CDS_Name"] = top50["Ticker"].map(cds_dict).fillna("")
    matched = (top50["CDS_Name"] != "").sum()
    print(f"  {matched}/{len(top50)} ticker icin CDS adi bulundu.")
    return top50


def match_names(top50_df: pd.DataFrame, graph_nodes: list[str]) -> dict[str, str]:
    """Graf dugum adi -> ticker esleme (fuzzy)."""
    node_to_ticker: dict[str, str] = {}
    used: set[str] = set()
    log = []

    for _, row in top50_df.iterrows():
        ticker   = row["Ticker"]
        cds_name = str(row.get("CDS_Name","")).strip()
        co_name  = str(row.get("Company","")).strip()
        names    = [n for n in [cds_name, co_name] if n]

        best_node, best_score = None, 0.0
        for target in names:
            for node in graph_nodes:
                if node in used:
                    continue
                s = _sim(target, node)
                if s > best_score:
                    best_score, best_node = s, node

        if best_node and best_score >= FUZZY_THRESHOLD:
            node_to_ticker[best_node] = ticker
            used.add(best_node)
            log.append((ticker, best_score, best_node, names[0] if names else ""))
        else:
            log.append((ticker, best_score, "NO_MATCH", names[0] if names else ""))

    matched = sum(1 for *_, n, _ in log if n != "NO_MATCH")
    print(f"\n  Esleme: {matched}/{len(top50_df)} ticker eslesti\n")
    print(f"  {'Ticker':<8} {'Skor':>6}  Graf dugumu                       CDS adi")
    print(f"  {'-'*80}")
    for ticker, score, node, cname in sorted(log):
        st = "OK" if node != "NO_MATCH" else "!!"
        print(f"  [{st}] {ticker:<8} {score:>6.3f}  {str(node)[:35]:<35}  {cname[:25]}")
    return node_to_ticker


def build_ticker_graph(
    G: nx.DiGraph,
    node_to_ticker: dict[str, str],
    tickers: list[str],
) -> nx.DiGraph:
    ts = set(tickers)
    G_tk = nx.DiGraph()
    for s, c in G.edges():
        st = node_to_ticker.get(s)
        ct = node_to_ticker.get(c)
        if st and ct and st in ts and ct in ts:
            G_tk.add_edge(st, ct)
    print(f"  Ticker grafi: {G_tk.number_of_nodes()} dugum, "
          f"{G_tk.number_of_edges()} kenar")
    return G_tk


def save_relations(G_tk, tickers, out_sup, out_cus):
    ts = set(tickers)
    wb_s = Workbook(); ws_s = wb_s.active
    ws_s.title = "Suppliers"; ws_s["A1"] = "Ticker"
    wb_c = Workbook(); ws_c = wb_c.active
    ws_c.title = "Customers"; ws_c["A1"] = "Ticker"
    ns = nc = 0
    for i, t in enumerate(tickers, 2):
        sups = sorted(s for s in G_tk.predecessors(t) if s in ts) if t in G_tk else []
        cuss = sorted(c for c in G_tk.successors(t)   if c in ts) if t in G_tk else []
        ws_s.cell(i, 1, t)
        ws_c.cell(i, 1, t)
        for j, s in enumerate(sups, 2): ws_s.cell(i, j, s)
        for j, c in enumerate(cuss, 2): ws_c.cell(i, j, c)
        if sups: ns += 1
        if cuss: nc += 1
    wb_s.save(out_sup); wb_c.save(out_cus)
    print(f"  [OK] {out_sup.name}  (bagli: {ns}/{len(tickers)})")
    print(f"  [OK] {out_cus.name}  (bagli: {nc}/{len(tickers)})")


def main():
    print("=" * 55)
    print("  04_build_sc_relations.py")
    print("=" * 55)
    make_dirs()

    print("\n[1/4] Graf yukleniyor...")
    G = load_graph()

    print("\n[2/4] Top-50 + CDS adlari yukleniyor...")
    top50_df = load_top50()
    tickers  = top50_df["Ticker"].tolist()

    print("\n[3/4] Fuzzy name matching...")
    node_to_ticker = match_names(top50_df, list(G.nodes()))

    print("\n[4/4] Dosyalar kaydediliyor...")
    G_tk = build_ticker_graph(G, node_to_ticker, tickers)
    save_relations(G_tk, tickers, TOP50_SUPPLIER_XLSX, TOP50_CUSTOMER_XLSX)

    if G_tk.number_of_edges() > 0:
        print("\n  Baglanti ornekleri:")
        count = 0
        for t in tickers:
            if t in G_tk:
                s = list(G_tk.predecessors(t))
                c = list(G_tk.successors(t))
                if s or c:
                    print(f"    {t:<6} sup={s[:3]}  cus={c[:3]}")
                    count += 1
                    if count >= 8: break

    print("\n[DONE] 04_build_sc_relations.py tamamlandi.")


if __name__ == "__main__":
    main()
